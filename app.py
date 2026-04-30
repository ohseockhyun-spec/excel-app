import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from io import BytesIO

st.set_page_config(page_title="주간 물성", layout="wide")

TARGET_CELLS = {
    "두께": "C21",
    "투습도": "M18",
    "흡수도": "M19",
    "점착력": "F8",
    "인장강도": "F13",
}

if "mode" not in st.session_state:
    st.session_state.mode = "주간 마이티 물성"

if "uploader_key" not in st.session_state:
    st.session_state.uploader_key = 0


def reset_uploader():
    st.session_state.uploader_key += 1


def change_mode(new_mode):
    if st.session_state.mode != new_mode:
        st.session_state.mode = new_mode
        reset_uploader()


def clear_all_files():
    reset_uploader()


def format_value(item_name, value):
    if value is None:
        return value

    try:
        num = float(value)

        if item_name == "두께":
            num = round(num * 1000)
            return f"{int(num):,}"

        elif item_name in ["투습도", "흡수도", "점착력"]:
            num = round(num)
            return f"{int(num):,}"

        elif item_name == "인장강도":
            num = round(num, 2)
            return f"{num:.2f}"

        return value

    except Exception:
        return value


def read_excel_values(uploaded_file):
    try:
        wb = load_workbook(uploaded_file, data_only=True)

        # 첫 번째 시트 강제 읽기
        ws = wb.worksheets[0]

        result = {}
        for item_name, cell_addr in TARGET_CELLS.items():
            raw_value = ws[cell_addr].value
            result[item_name] = format_value(item_name, raw_value)

        return result, None

    except Exception as e:
        return None, str(e)


def make_vertical_table(results_dict):
    df = pd.DataFrame(results_dict)

    lot_row = pd.DataFrame(
        [list(results_dict.keys())],
        columns=df.columns,
        index=["LOT"]
    )
    df = pd.concat([lot_row, df])

    df.index.name = "항목"
    return df


def to_excel_download(df):
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="결과")
    output.seek(0)
    return output


col1, col2, col3 = st.columns([1, 1, 1])

with col1:
    if st.button("주간 마이티 물성", use_container_width=True):
        change_mode("주간 마이티 물성")

with col2:
    if st.button("주간 원단 물성", use_container_width=True):
        change_mode("주간 원단 물성")

with col3:
    if st.button("업로드 전체 삭제", use_container_width=True):
        clear_all_files()

mode = st.session_state.mode

st.title(mode)
st.write("여러 엑셀 파일에서 지정 셀 값을 읽어 세로형 표로 정리합니다.")

uploaded_files = st.file_uploader(
    "엑셀 파일 여러 개 업로드 (.xlsx, .xlsm)",
    type=["xlsx", "xlsm"],
    accept_multiple_files=True,
    key=f"file_uploader_{st.session_state.uploader_key}"
)

if uploaded_files:
    results = {}
    errors = []

    for file in uploaded_files:
        values, err = read_excel_values(file)

        if err:
            errors.append({"파일명": file.name, "오류내용": err})
        else:
            # 확장자 제거 + "_" 이후 전부 삭제
            lot_name = file.name.rsplit(".", 1)[0].split("_")[0]
            results[lot_name] = values

    if results:
        df = make_vertical_table(results)

        st.subheader("결과")
        st.dataframe(df, use_container_width=True)

        excel_file = to_excel_download(df)
        filename = (
            "주간_마이티_물성.xlsx"
            if mode == "주간 마이티 물성"
            else "주간_원단_물성.xlsx"
        )

        st.download_button(
            label="엑셀 다운로드",
            data=excel_file,
            file_name=filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        # LOT 목록 / 개수
        lot_list = list(results.keys())
        lot_string = ", ".join(lot_list)

        st.markdown("---")
        st.subheader("LOT 목록")
        st.write(lot_string)
        st.write(f"총 {len(lot_list)}개")

    if errors:
        st.subheader("오류 파일")
        st.dataframe(pd.DataFrame(errors), use_container_width=True)

else:
    st.info("엑셀 파일을 업로드하세요.")
